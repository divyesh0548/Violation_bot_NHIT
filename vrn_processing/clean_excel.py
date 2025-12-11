from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def remove_empty_columns(filename, sheet_name, header_row):
    wb = load_workbook(filename)
    ws = wb[sheet_name]

    max_col = ws.max_column
    cols_to_delete = []

    # Scan every column starting from header_row downward
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        header_value = ws[f"{col_letter}{header_row}"].value

        # If header is empty - mark for delete
        if header_value is None or str(header_value).strip() == "":
            cols_to_delete.append(col_idx)
            continue

        # Check if entire column below header is empty
        empty_below = True
        for row_idx in range(header_row + 1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value not in [None, ""]:
                empty_below = False
                break

        if empty_below:
            cols_to_delete.append(col_idx)

    # Delete columns from right to left (important)
    for col_idx in sorted(cols_to_delete, reverse=True):
        ws.delete_cols(col_idx, 1)

    output = filename.replace(".xlsx", "_cleaned.xlsx")
    wb.save(output)
    return output

# remove_empty_columns("yourfile_unmerged.xlsx", "Transaction_Report", 11)
