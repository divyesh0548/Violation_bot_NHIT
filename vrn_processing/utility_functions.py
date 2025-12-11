from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
from botocore.exceptions import NoCredentialsError, ClientError
import boto3
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib
import pandas as pd
import os
import re
import requests
import openpyxl
from io import BytesIO
import logging

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

AWS_S3_BUCKET_NAME = "snt-nhit-data"
AWS_REGION = "us-east-1"
AWS_ACCESS_KEY_ID="AKIA4MBPFHU6H4CPD7VV"
AWS_SECRET_ACCESS_KEY="Pc7AZGNonfzHo6SAJ1LSoxpBGrOSmqH6CCQ/sLQV"

# Email configuration (update with your SMTP settings)
SMTP_SERVER = "smtp.gmail.com"  # Change as needed
SMTP_PORT = 587
SENDER_EMAIL = "bots.admin@sharpandtannan.com"  # Your email
SENDER_PASSWORD = "uyun sooi jvic qjyb"  # Your email app password

s3_client = boto3.client(
                's3',
                aws_access_key_id=AWS_ACCESS_KEY_ID,
                aws_secret_access_key=AWS_SECRET_ACCESS_KEY,
                region_name=AWS_REGION
            )

# Function to upload file to S3
def upload_file_to_s3(file, entity_name, file_type, shift, filename=None, content_type=None):
    try:
        # Generate secure filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Handle both FileStorage objects and regular file handles
        if hasattr(file, 'filename') and file.filename:
            # FileStorage-like object (Flask upload)
            original_filename = secure_filename(file.filename)
            file_content_type = file.content_type if hasattr(file, 'content_type') else content_type
        else:
            # Regular file handle - use provided filename or default
            if filename:
                original_filename = secure_filename(filename)
            else:
                original_filename = "file.xlsx"
            file_content_type = content_type or 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        filename = f"{timestamp}_{original_filename}"
        
        # Construct S3 key (path) - all files saved under Output_folder
        s3_key = f"Output_folder/{entity_name}/{file_type}/{shift}/{filename}"
        
        # Upload file to S3
        if not s3_client:
            raise Exception("S3 client not initialized. Check AWS credentials.")
        
        s3_client.upload_fileobj(
            file,
            AWS_S3_BUCKET_NAME,
            s3_key,
            ExtraArgs={
                'ContentType': file_content_type,
                'ContentDisposition': 'inline'  # For viewing in browser
            }
        )
        
        # Generate S3 URL
        s3_url = f"https://{AWS_S3_BUCKET_NAME}.s3.{AWS_REGION}.amazonaws.com/{s3_key}"
        
        return s3_url
        
    except NoCredentialsError:
        print("AWS credentials not available")
        return None
    except ClientError as e:
        print(f"Error uploading to S3: {e}")
        return None
    except Exception as e:
        print(f"Unexpected error: {e}")
        return None


def format_entity_name(entity_name):
    """
    Format entity name: remove special characters and capitalize first letter of each word.
    Example: 'odhaki_paipkhar' -> 'Odhaki Paipkhar'
    """
    # Split by special characters (underscore, dash, etc.) and filter out empty strings
    words = re.split(r'[_\-\s]+', entity_name)
    words = [word for word in words if word]  # Remove empty strings
    
    # Capitalize first letter of each word
    formatted_words = [word.capitalize() for word in words]
    
    # Join with spaces
    return ' '.join(formatted_words)


def check_excel_rows(file_url: str, timeout: int = 30) -> bool:
    excel_file = None
    workbook = None
    
    try:
        logger.info(f"Downloading file from: {file_url}")
        
        # Download the file
        response = requests.get(file_url, timeout=timeout)
        response.raise_for_status()  # Raise exception for bad status codes
        
        logger.info(f"File downloaded successfully. Size: {len(response.content)} bytes")
        
        # Load the Excel file into memory
        excel_file = BytesIO(response.content)
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        
        logger.info(f"Excel file loaded. Worksheets: {workbook.sheetnames}")
        
        # Get the active worksheet (first sheet by default)
        worksheet = workbook.active
        
        # Count rows with data (excluding empty rows)
        max_row = worksheet.max_row
        logger.info(f"Maximum row count: {max_row}")
        
        # Return True if rows > 1, False otherwise
        result = max_row > 1
        logger.info(f"Result: {result} (rows: {max_row})")
        
        return result
        
    except requests.exceptions.Timeout:
        logger.error(f"Request timeout while downloading: {file_url}")
        return False
    except requests.exceptions.ConnectionError:
        logger.error(f"Connection error while downloading: {file_url}")
        return False
    except requests.exceptions.RequestException as e:
        logger.error(f"Request error: {str(e)}")
        return False
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        return False
    finally:
        # Clean up - delete/close the Excel file and buffer
        try:
            if workbook:
                workbook.close()
                logger.info("Workbook closed successfully")
        except Exception as e:
            logger.warning(f"Error closing workbook: {str(e)}")
        
        try:
            if excel_file:
                excel_file.close()
                logger.info("File buffer cleared from memory")
        except Exception as e:
            logger.warning(f"Error clearing file buffer: {str(e)}")


def send_email_to_plaza(email_data):
    try:
        # Extract data from email_data dictionary
        recipient_emails = email_data.get('emails', [])
        plaza_name_raw = email_data.get('entity_name', '')
        shift_name = email_data.get('shift_name', '')
        submission_date = email_data.get('date', '')
        file_link = email_data.get('file_link', '')
        record_id = email_data.get('record_id', '')
        
        # Format plaza name: remove special characters and capitalize first letter of each word
        plaza_name = format_entity_name(plaza_name_raw)
        
        if not recipient_emails:
            print(f" No recipient emails found for Record ID {record_id}")
            return False
        print(f"  Original Recipients: {', '.join(recipient_emails)}")
    
        cc_emails = [
            "jyotibaghel@nhit.co.in",
            "vijaykumar@nhit.co.in",
            "rahulagrawal@nhit.co.in",
            "dhanshree.mahajan@sharpandtannan.com"
        ]
        
        # Check if file has violations using check_excel_rows
        has_violations = False
        if file_link:
            print(f"  Checking file for violations: {file_link}")
            has_violations = check_excel_rows(file_link)
            print(f"  File check result: {'Violations found' if has_violations else 'No violations found'}")
        
        # Create message
        msg = MIMEMultipart()
        msg['From'] = SENDER_EMAIL
        msg['To'] = ', '.join(recipient_emails)  # Handle multiple recipients
        msg['Cc'] = ', '.join(cc_emails)
        msg['Subject'] = f"{plaza_name} - Output file for {shift_name} shift"
        
        # Email body - conditional based on violations found
        if has_violations:
            # Default body with file link when violations are found
            body = f"""
Dear {plaza_name} Team,

Exceptional transactions are attached. Based on these, the validating teams may raise violations after verifying the AVCC image or vehicle image.

Shift Details:
- Date: {submission_date}
- Shift: {shift_name}

File Link: {file_link}

Best regards,
Sharp and Tannan Associates
            """
        else:
            # Body without file link when no violations found
            body = f"""
Dear {plaza_name} Team,

We have reviewed the data for the shift and found no exceptional transactions.

Shift Details:
- Date: {submission_date}
- Shift: {shift_name}

Best regards,
Sharp and Tannan Associates
            """
        
        msg.attach(MIMEText(body, 'plain'))
        
        # Combine To and CC for actual sending
        all_recipients = recipient_emails + cc_emails
        
        # Send email
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(SENDER_EMAIL, SENDER_PASSWORD)
        server.sendmail(SENDER_EMAIL, all_recipients, msg.as_string())
        server.quit()
        
        print(f"   Email sent to {plaza_name} (Record ID: {record_id})")
        print(f"   Recipients: {', '.join(recipient_emails)}")
        print(f"   CC: {', '.join(cc_emails)}")
        print(f"   Shift: {shift_name} | Date: {submission_date}")
        return True
    except Exception as e:
        # Use formatted name if available, otherwise format raw name for error message
        try:
            error_plaza_name = plaza_name if 'plaza_name' in locals() else format_entity_name(plaza_name_raw)
        except:
            error_plaza_name = email_data.get('entity_name', 'Unknown')
        print(f" Error sending email to {error_plaza_name}: {e}")
        return False
    
    
    
def read_csv_with_header_detection(path):
    header_keywords = [
        'Transaction ID', 'Date & Time', 'Veh Reg No.', 'TC Class', 'Lane No',
        'VEH REG NO', 'TRANSACTION NO', "TC_VEH_REG_NO", "PLAZA_NAME",
        'MVC_TLC_CLASS', 'MVC_TLC_MOP', 'MVC MOP', 'MVC',
        'Veh Reg Num', 'PaymentType', 'VehicleNumber'
    ]

    # ---- Step 1: Read raw lines
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        lines = f.readlines()

    detected_header_index = None

    for i, line in enumerate(lines):
        upper_line = line.upper()
        keyword_count = sum(1 for kw in header_keywords if kw.upper() in upper_line)

        if keyword_count >= 2:
            detected_header_index = i
            break

    if detected_header_index is None:
        raise Exception("Header could not be detected")

    print(f"  [INFO] Header detected at line: {detected_header_index+1}")

    # ---- Step 2: Trim everything ABOVE header
    trimmed = lines[detected_header_index:]

    # ---- Step 3: Save temporarily and read again
    temp_clean = path + "_clean.csv"
    with open(temp_clean, "w", encoding="utf-8") as f:
        f.writelines(trimmed)

    print("  [INFO] Re-reading trimmed CSV...")

    # ---- Step 4: Try flexible read
    try:
        df = pd.read_csv(temp_clean, engine="python", sep=None)
    except:
        df = pd.read_csv(temp_clean, engine="python", on_bad_lines="skip")

    # Clean temp file
    os.remove(temp_clean)

    return df





#To delete an object from S3 cloud.
# key_for_delete = "Output_folder/testing_entity/VRN/8_4/vrn_file_validated_20251122_Shift-8_4.xlsx"
# s3_client.delete_object(Bucket=AWS_S3_BUCKET_NAME, Key=key_for_delete)