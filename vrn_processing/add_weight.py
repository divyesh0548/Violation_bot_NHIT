import pandas as pd
from openpyxl import load_workbook


def add_weights_to_vrn_file(df_found, df_not_found_with_weights, vrn_file, sheet_name='VRNReport', header_index=1):
    print("=" * 80)
    print("STEP 1: Combining weight data from both dataframes...")
    print("=" * 80)
    
    # Validate columns
    if 'Veh Reg No.' not in df_found.columns or 'Weight' not in df_found.columns:
        raise ValueError("df_found must have 'Veh Reg No.' and 'Weight' columns")
    
    if 'Veh Reg No.' not in df_not_found_with_weights.columns or 'Weight' not in df_not_found_with_weights.columns:
        raise ValueError("df_not_found_with_weights must have 'Veh Reg No.' and 'Weight' columns")
    
    print("[OK] Both dataframes have correct columns")
    
    # Combine both DataFrames
    df_weights = pd.concat([df_found, df_not_found_with_weights], ignore_index=True)
    
    # Standardize vehicle numbers (uppercase, strip whitespace)
    df_weights['Veh Reg No.'] = df_weights['Veh Reg No.'].astype(str).str.strip().str.upper()
    
    # Remove duplicates (in case any exist), keeping first occurrence
    df_weights = df_weights.drop_duplicates(subset=['Veh Reg No.'], keep='first')
    
    print(f"Total unique vehicle-weight pairs: {len(df_weights)}")
    print(f"Sample data:")
    print(df_weights.head(10))
    
    # =============================================================================
    # STEP 2: Read the original VRN file
    # =============================================================================
    print("\n" + "=" * 80)
    print("STEP 2: Reading original VRN file...")
    print("=" * 80)
    
    # Read the Excel file with header at row 11
    df_vrn = pd.read_excel(vrn_file, sheet_name=sheet_name, header=header_index)
    
    print(f"Total records in VRN file: {len(df_vrn)}")
    
    # Standardize vehicle numbers in VRN file
    df_vrn['Veh Reg No.'] = df_vrn['Veh Reg No.'].astype(str).str.strip().str.upper()
    
    # =============================================================================
    # STEP 3: Merge weight data with VRN file
    # =============================================================================
    print("\n" + "=" * 80)
    print("STEP 3: Merging weight data with VRN file...")
    print("=" * 80)
    
    # Perform left join to add weight column
    df_vrn_updated = df_vrn.merge(
        df_weights[['Veh Reg No.', 'Weight']],
        on='Veh Reg No.',
        how='left'
    )
    
    # Count how many records got weight data
    records_with_weight = df_vrn_updated['Weight'].notna().sum()
    records_without_weight = df_vrn_updated['Weight'].isna().sum()
    
    print(f"Records matched with weight data: {records_with_weight}")
    print(f"Records without weight data: {records_without_weight}")
    
    # =============================================================================
    # STEP 4: Save the updated VRN file
    # =============================================================================
    print("\n" + "=" * 80)
    print("STEP 4: Saving updated VRN file...")
    print("=" * 80)
    
    output_file = 'vrn_file_updated.xlsx'
    
    # Load the original workbook to preserve formatting and metadata rows
    wb = load_workbook(vrn_file)
    ws = wb[sheet_name]
    
    # Convert zero-based header index (pandas style) to Excel row number (1-based)
    header_row = header_index + 1
    
    # Find the 'Veh Reg No.' column
    veh_reg_col = None
    for col_idx, cell in enumerate(ws[header_row], start=1):
        if cell.value == 'Veh Reg No.':
            veh_reg_col = col_idx
            break
    
    if veh_reg_col is None:
        raise ValueError("Could not find 'Veh Reg No.' column in VRN file")
    
    # Add 'Weight' column header next to last column
    last_col = ws.max_column
    weight_col = last_col + 1
    ws.cell(row=header_row, column=weight_col, value='Weight')
    
    print(f"[OK] Added 'Weight' column at position {weight_col}")
    
    # Create a lookup dictionary for faster access
    weight_lookup = dict(zip(df_weights['Veh Reg No.'], df_weights['Weight']))
    
    # Update weight values starting from row 13 (data starts after header)
    records_updated = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        veh_reg_no = ws.cell(row=row_idx, column=veh_reg_col).value
        
        if veh_reg_no:
            # Standardize the vehicle number
            veh_reg_no_std = str(veh_reg_no).strip().upper()
            
            # Look up weight
            if veh_reg_no_std in weight_lookup:
                weight_value = weight_lookup[veh_reg_no_std]
                ws.cell(row=row_idx, column=weight_col, value=weight_value)
                records_updated += 1
    
    print(f"[OK] Updated {records_updated} records with weight data")
    
    # Save the workbook
    wb.save(output_file)
    
    print(f"[OK] Updated VRN file saved as '{output_file}'")
    
    # =============================================================================
    # STEP 5: Verification
    # =============================================================================
    print("\n" + "=" * 80)
    print("STEP 5: Verification...")
    print("=" * 80)
    
    # Read the updated file to verify
    df_verify = pd.read_excel(output_file, sheet_name=sheet_name, header=header_index)
    
    if 'Weight' in df_verify.columns:
        print("[OK] 'Weight' column successfully added")
        print(f"Records with weight data: {df_verify['Weight'].notna().sum()}")
        print(f"Records without weight data: {df_verify['Weight'].isna().sum()}")
        print("\nSample data with weights:")
        print(df_verify[['Veh Reg No.', 'Weight']].head(10))
    else:
        print("[Error]: 'Weight' column not found in updated file")
    
    # =============================================================================
    # SUMMARY
    # =============================================================================
    print("\n" + "=" * 80)
    print("SUMMARY")
    print("=" * 80)
    print(f"Input file: {vrn_file}")
    print(f"Output file: {output_file}")
    print(f"\nTotal VRN records: {len(df_vrn)}")
    print(f"Vehicle-weight pairs available: {len(df_weights)}")
    print(f"Records updated with weight: {records_updated}")
    print(f"Records without weight: {len(df_vrn) - records_updated}")
    print("=" * 80)
    print("[OK] Process completed successfully!")
    print("=" * 80)
    
    return output_file


# Example usage when running standalone
if __name__ == "__main__":
    # Load dataframes from Excel for testing
    results_file = 'vehicle_check_results.xlsx'
    
    df_found = pd.read_excel(results_file, sheet_name='Found')
    df_not_found = pd.read_excel(results_file, sheet_name='Not Found')
    
    print(f"Loaded {len(df_found)} found vehicles")
    print(f"Loaded {len(df_not_found)} not found vehicles")
    
    # Add weights to VRN file
    output_file = add_weights_to_vrn_file(
        df_found=df_found,
        df_not_found_with_weights=df_not_found,
        vrn_file='vrn_file.xlsx',
        sheet_name='VRNReport',
        header_index=11
    )
    
    print(f"\n[OK] Final output: {output_file}")
