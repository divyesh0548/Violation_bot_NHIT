import pandas as pd
import os
from openpyxl import load_workbook


# Mapping configurations
WEIGHT_RANGE_INDEXES = [
    (7500, 1, "Car"),
    (12000, 2, "LCV"),
    (18500, 3, "Truck"),
    (31000, 4, "Truck3X"),
    (60000, 5, "MAV"),
]

TC_CLASS_INDEX_MAP = {
    'car': 1,
    'lcv': 2,
    'minibus': 2,
    'lcv/mini bus': 2,
    'truck': 3,
    'truck/bus': 3,
    'bus/truck': 3,
    'bus': 3,
    'bus 2 axle': 3,
    'bus-2 axle': 3,
    'truck 2 axle': 3,
    'truck-2 axle': 3,
    'bus 3 Ax': 4,
    "bus/truck 3 axle": 4,
    "bus-3 axle" : 4,
    "truck 3 axle" : 4,
    "truck-3 axle" : 4,
    'truck3x': 4,
    'truck-3ax': 4,
    'mav': 5,
    'hcm/mav' : 5,
    'mav (4 to 6 axle)': 5,
    'mav 4 ax':5,
    'mav 4 axle': 5,
    'mav 4 axle -2t' : 5,
    'mav 5 ax': 5,
    'mav 5 axle': 5,
    'mav 5 axle - 5t' : 5,
    'mav 5 axle - 5t1' : 5,
    'mav 5 axle - 5t4' : 5,
    'mav 5 axle -2t' : 5,
    'mav 5 axle -3t' : 5,
    'mav 6 ax' : 5,
    'mav 6 axle' : 5,
    'mav 6 axle - 6t1' : 5,
    'truck 4-6 axle' : 5,
    'osv': 6,
}

WEIGHT_CLASS_INDEX_MAP = {
    'car': 1,
    'lcv': 2,
    'truck': 3,
    'truck3x': 4,
    'mav': 5,
    'osv': 6,
}

TC_CLASS_COLUMNS = ['TC Class', 'TC_Class', 'TcClass', 'MVC Class', 'Operator Class', 'MVC_TLC_CLASS', 'MVC', 'MVC (TLC CLASS)', 'VEH CLASS', 'Veh Class']

def get_vehicle_class_from_weight(weight):
    try:
        # Convert weight to numeric
        weight_value = float(weight)
        
        # Apply weight class logic
        if weight_value < 100:
            return "NA"
        for upper_bound, _, label in WEIGHT_RANGE_INDEXES:
            if weight_value <= upper_bound:
                return label
        if weight_value > 60000:
            return "OSV"
        return "NA"
    
    except (ValueError, TypeError):
        # If weight is not a valid number (N/A, Error, etc.)
        return "Unknown"


def validate_vehicle_classes(vrn_file_path, sheet_name='VRNReport', header_index=1):
    
    # =============================================================================
    # STEP 1: Read VRN file
    # =============================================================================
    print("=" * 80)
    print("STEP 1: Reading VRN file...")
    print("=" * 80)
    
    df_vrn = pd.read_excel(vrn_file_path, sheet_name=sheet_name, header=header_index)
    
    print(f"Total records in VRN file: {len(df_vrn)}")
    
    # Check required columns
    if 'Weight' not in df_vrn.columns:
        raise ValueError("VRN file must have 'Weight' column. Please run add_weights_to_vrn_file first.")
    
    tc_class_col = next((col for col in TC_CLASS_COLUMNS if col in df_vrn.columns), None)
    if tc_class_col is None:
        raise ValueError(f"VRN file must have one of these columns: {TC_CLASS_COLUMNS}")
    
    print(f"[OK] Required columns found: 'Weight' and '{tc_class_col}'")
    
    # =============================================================================
    # STEP 2: Apply weight-based classification
    # =============================================================================
    print("\n" + "=" * 80)
    print("STEP 2: Applying weight-based vehicle classification...")
    print("=" * 80)
    
    # Create new column with weight-based class
    df_vrn['Weight_Based_Class'] = df_vrn['Weight'].apply(get_vehicle_class_from_weight)
    
    print("[OK] Weight-based classification applied")
    print(f"\nClassification distribution:")
    print(df_vrn['Weight_Based_Class'].value_counts())
    
    # =============================================================================
    # STEP 3: Compare with TC Class
    # =============================================================================
    print("\n" + "=" * 80)
    print("STEP 3: Comparing TC Class with Weight-based Class...")
    print("=" * 80)
    
    # Normalize both classes to lowercase for comparison and mapping

    df_vrn['TC_Class_Lower'] = df_vrn[tc_class_col].astype(str).str.lower().str.strip()
    df_vrn['Weight_Class_Lower'] = df_vrn['Weight_Based_Class'].astype(str).str.lower().str.strip()
    
    # Map to indexes
    df_vrn['TC_class_index'] = df_vrn['TC_Class_Lower'].map(TC_CLASS_INDEX_MAP)
    df_vrn['Correct_index'] = df_vrn['Weight_Class_Lower'].map(WEIGHT_CLASS_INDEX_MAP)
    
    # Result column: difference between indexes
    df_vrn['result'] = df_vrn['Correct_index'] - df_vrn['TC_class_index']
    
    # class_accuracy based on result value
    df_vrn['class_accuracy'] = df_vrn['result'].apply(
        lambda x: 'TRUE' if pd.notna(x) and x == 0 else 'FALSE'
    )
    
    # Drop temporary lowercase columns
    df_vrn = df_vrn.drop(columns=['TC_Class_Lower', 'Weight_Class_Lower'])
    
    # Calculate accuracy statistics
    total_records = len(df_vrn)
    records_with_weight = df_vrn['Weight'].notna().sum()
    accurate_count = (df_vrn['class_accuracy'] == 'TRUE').sum()
    inaccurate_count = total_records - accurate_count
    
    # Count unknown classifications (where weight was invalid)
    unknown_count = (df_vrn['Weight_Based_Class'] == 'Unknown').sum()
    
    print(f"Total records: {total_records}")
    print(f"Records with valid weight: {records_with_weight}")
    print(f"Records with unknown weight: {unknown_count}")
    print(f"Accurate classifications: {accurate_count}")
    print(f"Inaccurate classifications: {inaccurate_count}")
    print(f"TC class index distribution:\n{df_vrn['TC_class_index'].value_counts(dropna=False)}")
    print(f"Weight index distribution:\n{df_vrn['Correct_index'].value_counts(dropna=False)}")
    
    if records_with_weight > 0:
        accuracy_percentage = (accurate_count / total_records) * 100
        print(f"Overall accuracy: {accuracy_percentage:.2f}%")
    
    # =============================================================================
    # STEP 4: Save validated VRN file
    # =============================================================================
    print("\n" + "=" * 80)
    print("STEP 4: Saving validated VRN file...")
    print("=" * 80)
    
    output_file = 'vrn_file_validated.xlsx'
    
    # Load the original workbook to preserve formatting and metadata rows
    wb = load_workbook(vrn_file_path)
    ws = wb[sheet_name]
    
    # Add new column headers
    last_col = ws.max_column
    new_columns = [
        ('Weight_Based_Class', 'Weight_Based_Class'),
        ('class_accuracy', 'class_accuracy'),
        ('TC_class_index', 'TC_class_index'),
        ('Correct_index', 'Correct_index'),
        ('result', 'result'),
    ]
    
    header_row = header_index + 1
    for offset, (_, header_name) in enumerate(new_columns, start=1):
        ws.cell(row=header_row, column=last_col + offset, value=header_name)
    
    print("[OK] Added new columns:", ", ".join(name for _, name in new_columns))
    
    # Update the data rows
    for row_idx, (_, row_data) in enumerate(df_vrn.iterrows(), start=header_row + 1):
        for offset, (df_col, _) in enumerate(new_columns, start=1):
            ws.cell(row=row_idx, column=last_col + offset, value=row_data[df_col])
    
    # Save the workbook
    wb.save(output_file)
    
    print(f"[OK] Validated VRN file saved as '{output_file}'")
    
    # =============================================================================
    # STEP 5: Delete input file and rename output
    # =============================================================================
    print("\n" + "=" * 80)
    print("STEP 5: Replacing input file with validated file...")
    print("=" * 80)
    
    try:
        # Delete the input file
        os.remove(vrn_file_path)
        print(f"[OK] Deleted input file: {vrn_file_path}")
        
        # Rename output file to input file name
        os.rename(output_file, vrn_file_path)
        print(f"[OK] Renamed '{output_file}' to '{vrn_file_path}'")
        
        final_file = vrn_file_path
    
    except Exception as e:
        print(f"[Error] Error replacing file: {e}")
        print(f"Keeping both files: {vrn_file_path} and {output_file}")
        final_file = output_file
        
        
    # =============================================================================
    # STEP 6: Create filtered file where result > 0
    # =============================================================================
    print("\n" + "=" * 80)
    print("STEP 6: Creating filtered result file (result > 0 only)...")
    print("=" * 80)
    
    try:
        # Read the final file again
        df_final = pd.read_excel(final_file, sheet_name=sheet_name, header=header_index)
    
        # Remove rows where result <= 0
        df_filtered = df_final[df_final["result"] > 0].copy()
    
        # Save filtered file
        filtered_file = final_file.replace(".xlsx", "_filtered.xlsx")
        df_filtered.to_excel(filtered_file, index=False)
    
        print(f"[OK] Filtered file created: {filtered_file}")
    
    except Exception as e:
        print(f"[Error] Error while creating filtered file: {e}")
        filtered_file = None
    
    # =============================================================================
    # SUMMARY
    # =============================================================================
    print("\n" + "=" * 80)
    print("SUMMARY")
    print("=" * 80)
    print(f"Final VRN file: {final_file}")
    print(f"Total records: {total_records}")
    print(f"Records with valid weight: {records_with_weight}")
    # print(f"Accurate classifications: {accurate_count} ({accuracy_percentage:.2f}%)")
    print(f"Inaccurate classifications: {inaccurate_count}")
    print(f"Unknown classifications: {unknown_count}")
    print("=" * 80)
    print("[OK] Vehicle class validation completed successfully!")
    print("=" * 80)
    
    return final_file


# Example usage when running standalone
if __name__ == "__main__":
    vrn_file = 'vrn_file_updated.xlsx'
    
    final_file = validate_vehicle_classes(
        vrn_file_path=vrn_file,
        sheet_name='VRNReport'
    )
    
    print(f"\n[OK] Final validated file: {final_file}")
