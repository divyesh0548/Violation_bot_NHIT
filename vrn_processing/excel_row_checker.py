"""
Excel File Row Counter Module
Downloads Excel file from URL and checks row count
"""

import requests
import openpyxl
from io import BytesIO
import logging

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def check_excel_rows(file_url: str, timeout: int = 30) -> bool:
    excel_file = None
    workbook = None
    
    try:
        logger.info(f"Downloading file from: {file_url}")
        
        # Download the file
        response = requests.get(file_url, timeout=timeout)
        response.raise_for_status()  # Raise exception for bad status codes
        
        logger.info(f"File downloaded successfully. Size: {len(response.content)} bytes")
        
        # Load the Excel file into memory
        excel_file = BytesIO(response.content)
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        
        logger.info(f"Excel file loaded. Worksheets: {workbook.sheetnames}")
        
        # Get the active worksheet (first sheet by default)
        worksheet = workbook.active
        
        # Count rows with data (excluding empty rows)
        max_row = worksheet.max_row
        logger.info(f"Maximum row count: {max_row}")
        
        # Return True if rows > 1, False otherwise
        result = max_row > 1
        logger.info(f"Result: {result} (rows: {max_row})")
        
        return result
        
    except requests.exceptions.Timeout:
        logger.error(f"Request timeout while downloading: {file_url}")
        return False
    except requests.exceptions.ConnectionError:
        logger.error(f"Connection error while downloading: {file_url}")
        return False
    except requests.exceptions.RequestException as e:
        logger.error(f"Request error: {str(e)}")
        return False
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        return False
    finally:
        # Clean up - delete/close the Excel file and buffer
        try:
            if workbook:
                workbook.close()
                logger.info("Workbook closed successfully")
        except Exception as e:
            logger.warning(f"Error closing workbook: {str(e)}")
        
        try:
            if excel_file:
                excel_file.close()
                logger.info("File buffer cleared from memory")
        except Exception as e:
            logger.warning(f"Error clearing file buffer: {str(e)}")


if __name__ == "__main__":
    # Example usage
    file_link = "https://snt-nhit-data.s3.us-east-1.amazonaws.com/raksha/VRN/12_8/20251208_024151_VRN_TRANSACTION_DETAIL_REPORT_A.xlsx"
    
    result = check_excel_rows(file_link)
    print(f"\nFile has more than 1 row: {result}")
