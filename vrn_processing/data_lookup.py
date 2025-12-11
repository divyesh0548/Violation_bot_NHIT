import pandas as pd
import psycopg2
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# --- Database Configuration ---
DB_HOST = "db-1.c2n44a20y9k5.us-east-1.rds.amazonaws.com"
DB_PORT = 5432
DB_NAME = "nhit"
DB_USER = "postgres"
DB_PASSWORD = "postgres1234"

# Column name fallbacks
VEHICLE_NUMBER_COLUMNS = [
    'Veh Reg No.', 'Veh Reg No', 'Vehicle Number', 'VehicleNumber', 'Vehicle No.', 'Vehicle No', 'TC_VEH_REG_NO',
    'Vehicle Registration Number', 'VRN', 'Registration No.', 'TC VEH REG NO', 
    'Unique Vehicle Number', 'Vehicle_Number', 'Vehicle_No', 'Veh_Reg_No', 'VEH REG NO', 'Veh Reg Num']

TC_CLASS_COLUMNS = ['TC Class', 'TC_Class', 'TcClass', 'MVC Class', 'Operator Class', 'MVC_TLC_CLASS', 'MVC', 'MVC (TLC CLASS)', 'VEH CLASS', 'Veh Class']
PAYMENT_METHOD_COLUMNS = ['Payment Method', 'Payment_Method', 'PAYMENT METHOD', 'PAYMENT_TYPE', 'PaymentType', 'MVC MOP', 'PAYMENT TYPE', 'MVC (TLC MOP)', 'Payment Mode']
PAYMENT_METHOD_COLUMNS_SECONDARY = ['Journey Type']

def get_db_connection():
    """Create and return a PostgreSQL database connection"""
    try:
        conn = psycopg2.connect(
            host=DB_HOST,
            port=DB_PORT,
            database=DB_NAME,
            user=DB_USER,
            password=DB_PASSWORD
        )
        return conn
    except psycopg2.Error as e:
        print(f"Error connecting to database: {e}")
        return None


def process_vehicle_data(excel_file='vrn_file.xlsx', sheet_name='VRNReport', header_index=1):

    print("=" * 80)
    print("STEP 1: Reading Excel file...")
    print("=" * 80)
    
    print(f"Header row index: {header_index}")
    
    df = pd.read_excel(excel_file, sheet_name=sheet_name, header=header_index)
    
    # Normalize vehicle number column name
    df_normalized = {col.strip().lower(): col for col in df.columns}
    veh_col_key = next(
    (col.strip().lower() for col in VEHICLE_NUMBER_COLUMNS
     if col.strip().lower() in df_normalized),
    None
    )
    
    # veh_col = next((col for col in VEHICLE_NUMBER_COLUMNS if col in df.columns), None)
    
    print(f"Vehicle number column key: {veh_col_key}")
    
    if veh_col_key is None:
        print(df.columns.tolist())
        raise ValueError("Could not find vehicle number column in VRN file")
    print(df.columns.tolist())
    
    veh_col = df_normalized[veh_col_key]
    
    print(f"Vehicle columns {veh_col}")
    
    if veh_col is None:
        raise ValueError("Could not find vehicle number column in VRN file")
    if veh_col != 'Veh Reg No.':
        df = df.rename(columns={veh_col: 'Veh Reg No.'})
    
    print(f"Total records loaded: {len(df)}")
    
    # =============================================================================
    # STEP 2: Remove duplicate vehicle numbers
    # =============================================================================
    print("\n" + "=" * 80)
    print("STEP 2: Removing duplicate vehicle numbers...")
    print("=" * 80)
    
    print(f"Records before removing duplicates: {len(df)}")
    
    # Find duplicates before removing
    duplicates = df[df.duplicated(subset=['Veh Reg No.'], keep=False)]
    duplicate_count = len(duplicates)
    unique_duplicate_vehicles = duplicates['Veh Reg No.'].nunique()
    
    print(f"Total duplicate records found: {duplicate_count}")
    print(f"Unique vehicles with duplicates: {unique_duplicate_vehicles}")
    
    # =============================================================================
    # STEP 2.1: Save deduplicated data back to Excel file (FAST METHOD)
    # =============================================================================
    print("\n" + "=" * 80)
    print("STEP 2.1: Saving deduplicated data back to original file...")
    print("=" * 80)
    
    # Write deduplicated data back while preserving rows above header
    header_row = header_index + 1  # convert zero-based index to Excel row number
    data_start_row = header_row + 1
    
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]
    
    # Ensure header row stays in sync with DataFrame columns
    for col_idx, column_name in enumerate(df.columns, start=1):
        ws.cell(row=header_row, column=col_idx, value=column_name)
    
    # Clear existing data rows below the header
    existing_data_rows = ws.max_row - header_row
    if existing_data_rows > 0:
        ws.delete_rows(data_start_row, existing_data_rows)
    
    # Write new data rows
    for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), start=data_start_row):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(excel_file)
    
    print(f"[OK] Deduplicated data saved back to '{excel_file}'")
    print(f"[OK] Original file updated with {len(df)} unique vehicle records")
    print(f"[OK] Removed {duplicate_count} duplicate records")
    
    # =============================================================================
    # STEP 3: Apply TC Class filters
    # =============================================================================
    print("\n" + "=" * 80)
    print("STEP 3: Applying TC Class filters (removing specific classes)...")
    print("=" * 80)
    
    print(f"Records before TC Class filter: {len(df)}")
    
    # Determine which column to use for TC class filtering
    tc_class_col = next((col for col in TC_CLASS_COLUMNS if col in df.columns), None)
    excluded_classes = ['CarJeep', 'Carjeep', r'CAR\JEEP', r'CAR/JEEP', r'CAR/JEEP/VAN', 'CJV', 'OSV', 'Auto', 'Blanks', 'Tractor','TRACTOR', '3WHEELER', '2WHEELER', 'THREE WHEELER', 'Three Wheeler', 'TWO WHEELER','UNDEFINED', 'TRACTOR - TRAILER', 'NA', 'N/A', 'N/A', '-', ' - ', 'NOCLASS', 'No Class']
    
    if tc_class_col:
        df = df[~df[tc_class_col].isin(excluded_classes)]
        print(f"[OK] Applied TC Class filter using column '{tc_class_col}'")
    else:
        print("[Warning] Warning: No TC Class column found from list; skipping TC filter")
    
    print(f"Records after TC Class filter: {len(df)}")
    
    # =============================================================================
    # STEP 4: Apply Payment Method filter
    # =============================================================================
    print("\n" + "=" * 80)
    print("STEP 4: Applying Payment Method filter (removing 'Exempt')...")
    print("=" * 80)
    
    print(f"Records before Payment Method filter: {len(df)}")
    
    # payment_col_primary = next((col for col in PAYMENT_METHOD_COLUMNS if col in df.columns), None)
    payment_col_primary = next((col for col in df.columns if any(
    col == variant or 
    col.replace('\n', ' ').replace('\r', ' ').strip() == variant.replace('\n', ' ').strip()
    for variant in PAYMENT_METHOD_COLUMNS
    )), None)
    payment_col_secondary = next((col for col in PAYMENT_METHOD_COLUMNS_SECONDARY if col in df.columns), None)
    exclude_values = ['Exempt','EXEMPT','EXEMPTED', 'EXEMPT-CASH ', 'EXEMPT-CASH', 'EXEMPT-NA ', 'EXEMPT-NA', 'Card', 'CASH-CASH', 'CASH-e-Wallet', 'CASH', 'Cash' , 'CASH-DEBIT/CREDIT CARD', 'CASH-STATIC UPI', 'UPI', 'CASH-UPI', 'Static UPI', 'EXEMPT-NA','RUNTHROUGH-NA', 'RUNTHROUGH-NA ', 'VIOLATION', 'Violation' 'ETC_PENALTY_CASH', 'Fleet', 'Fleet ', 'ETC_PENALTY_CASH', 'No Fee']  # List of strings to exclude
    include_values = ['CCH Media Trip']
    
    if payment_col_primary:
        # Primary column found - exclude specific values
        df = df[~df[payment_col_primary].isin(exclude_values)]

    elif payment_col_secondary:
        # Secondary column found - include only specific values
        df = df[df[payment_col_secondary].isin(include_values)]
        
    else:
        print("Warning: No payment method column found in either list")

    
    print(f"Records after Payment Method filter: {len(df)}")
    
    # =============================================================================
    # STEP 5: Extract and clean vehicle numbers
    # =============================================================================
    print("\n" + "=" * 80)
    print("STEP 5: Extracting vehicle numbers...")
    print("=" * 80)
    
    # Standardize vehicle numbers (trim whitespace, convert to uppercase)
    df['Veh Reg No.'] = df['Veh Reg No.'].astype(str).str.strip().str.upper()
    
    # Get unique vehicle numbers
    vehicle_numbers = df['Veh Reg No.'].unique().tolist()
    
    print(f"Total unique vehicle numbers after all filters: {len(vehicle_numbers)}")
    print(f"Sample vehicle numbers: {vehicle_numbers[:5]}")
    
    # =============================================================================
    # STEP 6: Query database
    # =============================================================================
    print("\n" + "=" * 80)
    print("STEP 6: Querying database...")
    print("=" * 80)
    
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            
            # Query to fetch matching vehicle numbers and weight
            query = """
                SELECT "Unique Vehicle Number", weight
                FROM checkpostmaster
                WHERE "Unique Vehicle Number" = ANY(%s)
            """
            
            print("Executing database query...")
            cursor.execute(query, (vehicle_numbers,))
            results = cursor.fetchall()
            
            # Convert to DataFrame
            df_found = pd.DataFrame(results, columns=['Veh Reg No.', 'Weight'])
            
            if len(df_found) > 0:
                # Standardize vehicle numbers in database results
                df_found['Veh Reg No.'] = df_found['Veh Reg No.'].astype(str).str.strip().str.upper()
                df_found = df_found.drop_duplicates(subset=['Veh Reg No.'], keep='first')
            
            cursor.close()
            conn.close()
            
            print(f"[OK] Database query completed successfully")
            print(f"Vehicle numbers found in database: {len(df_found)}")
            
        except psycopg2.Error as e:
            print(f"[Error] Error executing query: {e}")
            df_found = pd.DataFrame(columns=['Veh Reg No.', 'Weight'])
    else:
        print("[Error] Failed to connect to database")
        df_found = pd.DataFrame(columns=['Veh Reg No.', 'Weight'])
    
    # =============================================================================
    # STEP 7: Separate found and not found
    # =============================================================================
    print("\n" + "=" * 80)
    print("STEP 7: Separating found and not found vehicle numbers...")
    print("=" * 80)
    
    # Create DataFrame with all filtered vehicle numbers
    df_all_vehicles = pd.DataFrame({'Veh Reg No.': vehicle_numbers})
    
    # Not found vehicles (vehicles not in database)
    df_not_found = df_all_vehicles[
        ~df_all_vehicles['Veh Reg No.'].isin(df_found['Veh Reg No.'])
    ].copy()
    
    print(f"Found in database: {len(df_found)} vehicles")
    print(f"Not found in database (before capacity check): {len(df_not_found)} vehicles")
    
    # =============================================================================
    # STEP 7.0: Remove vehicles found in capacity_vehicle_numbers table
    # =============================================================================
    print("\n" + "=" * 80)
    print("STEP 7.0: Checking capacity_vehicle_numbers table...")
    print("=" * 80)
    
    if len(df_not_found) > 0:
        conn = get_db_connection()
        if conn:
            try:
                cursor = conn.cursor()
                
                # Get vehicle numbers from df_not_found
                not_found_vehicle_numbers = df_not_found['Veh Reg No.'].unique().tolist()
                
                # Query to fetch matching vehicle numbers from capacity_vehicle_numbers table
                capacity_query = """
                    SELECT "Unique Vehicle Number"
                    FROM capacity_vehicle_numbers
                    WHERE "Unique Vehicle Number" = ANY(%s)
                """
                
                print(f"Checking {len(not_found_vehicle_numbers)} vehicles against capacity_vehicle_numbers table...")
                cursor.execute(capacity_query, (not_found_vehicle_numbers,))
                capacity_results = cursor.fetchall()
                
                # Convert to list of vehicle numbers
                capacity_vehicle_numbers = [row[0] for row in capacity_results]
                
                if len(capacity_vehicle_numbers) > 0:
                    # Standardize vehicle numbers
                    capacity_vehicle_numbers = [str(v).strip().upper() for v in capacity_vehicle_numbers]
                    
                    # Remove vehicles found in capacity_vehicle_numbers from df_not_found
                    initial_count = len(df_not_found)
                    df_not_found = df_not_found[
                        ~df_not_found['Veh Reg No.'].isin(capacity_vehicle_numbers)
                    ].copy()
                    removed_count = initial_count - len(df_not_found)
                    
                    print(f"[OK] Removed {removed_count} vehicles found in capacity_vehicle_numbers table")
                else:
                    print(f"[OK] No vehicles found in capacity_vehicle_numbers table")
                
                cursor.close()
                conn.close()
                
            except psycopg2.Error as e:
                print(f"[Error] Error querying capacity_vehicle_numbers table: {e}")
        else:
            print("[Error] Failed to connect to database for capacity check")
    
    print(f"Not found in database (after capacity check): {len(df_not_found)} vehicles")
    
    # =============================================================================
    # STEP 7.1: Apply length validation to not found numbers
    # =============================================================================
    print("\n" + "=" * 80)
    print("STEP 7.1: Applying length validation to not found numbers...")
    print("=" * 80)
    
    # Calculate length of each vehicle number
    df_not_found['Length'] = df_not_found['Veh Reg No.'].str.len()
    
    print(f"Length distribution before filtering:")
    print(df_not_found['Length'].value_counts().sort_index())
    
    # Keep only vehicle numbers with length between 7 and 12 (inclusive)
    df_not_found_valid = df_not_found[
        (df_not_found['Length'] >= 7) & (df_not_found['Length'] <= 12)
    ].copy()
    
    # Count invalid entries
    invalid_too_short = len(df_not_found[df_not_found['Length'] < 7])
    invalid_too_long = len(df_not_found[df_not_found['Length'] > 12])
    
    print(f"\nRemoved {invalid_too_short} vehicle numbers (length < 7)")
    print(f"Removed {invalid_too_long} vehicle numbers (length > 12)")
    print(f"Valid not found vehicle numbers: {len(df_not_found_valid)}")
    
    # Drop the temporary 'Length' column
    df_not_found_valid = df_not_found_valid.drop(columns=['Length'])
    
    # Remove duplicate entries based on vehicle number
    df_not_found_valid = df_not_found_valid.drop_duplicates(subset=['Veh Reg No.'], keep='first')
    
    print(f"Not found in database (after length validation): {len(df_not_found_valid)} vehicles")
    
    # =============================================================================
    # SUMMARY
    # =============================================================================
    print("\n" + "=" * 80)
    print("SUMMARY")
    print("=" * 80)
    print(f"Total records processed: {len(df)}")
    print(f"Found in checkpostmaster: {len(df_found)}")
    print(f"Not found in checkpostmaster (valid length): {len(df_not_found_valid)}")
    print(f"Excluded from 'Not Found' (invalid length): {invalid_too_short + invalid_too_long}")
    print(f"  - Too short (< 7 characters): {invalid_too_short}")
    print(f"  - Too long (> 12 characters): {invalid_too_long}")
    print("=" * 80)
    print("[OK] Process completed successfully!")
    
    
    
    #Remove this code in deployment, only for testing when running this file directly
    output_file = 'vehicle_check_results.xlsx'
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_found.to_excel(writer, sheet_name='Found', index=False)
        df_not_found_valid.to_excel(writer, sheet_name='Not Found', index=False)

    
    # Return the two dataframes
    return df_found, df_not_found_valid


# Example usage (when running this file directly)
if __name__ == '__main__':
    df_found, df_not_found = process_vehicle_data('vrn_file.xlsx', 'VRNReport')
    
    # Optional: Save to Excel if running standalone
    output_file = 'vehicle_check_results.xlsx'
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_found.to_excel(writer, sheet_name='Found', index=False)
        df_not_found.to_excel(writer, sheet_name='Not Found', index=False)
    print(f"\n[OK] Results also saved to '{output_file}'")
