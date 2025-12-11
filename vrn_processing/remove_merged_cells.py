from openpyxl import load_workbook

wb = load_workbook("vrn_file.xlsx")
ws = wb.active

# Unmerge everything
for merged_range in list(ws.merged_cells):
    ws.unmerge_cells(range_string=str(merged_range))

wb.save("yourfile_unmerged.xlsx")
