def remove_duplicate_headers(vrn_file_path, sheet_name='VRNReport'):
    print("=" * 80)
    print("STEP 1: Reading VRN file and identifying duplicate headers...")
    print("=" * 80)
    
    # Read entire file without header to examine structure
    df_full = pd.read_excel(vrn_file_path, sheet_name=sheet_name, header=None)
    
    print(f"Total rows in file: {len(df_full)}")
    print(f"Sheet name: {sheet_name}")
    
    rows_to_remove = []
    
    # Detect duplicate header rows
    for i in range(len(df_full) - 1):
        current_row = df_full.iloc[i]
        next_row = df_full.iloc[i + 1]

        current_str = current_row.fillna('').astype(str)
        next_str = next_row.fillna('').astype(str)

        if current_str.equals(next_str):
            has_text = False
            for val in current_row:
                if pd.notna(val) and isinstance(val, str) and len(str(val).strip()) > 0:
                    has_text = True
                    break
            if has_text:
                rows_to_remove.append(i)
                print(f"Found duplicate at rows {i} and {i+1} - removing {i}")
                print(f"  Sample values: {current_row.iloc[0]}, {current_row.iloc[1]}, {current_row.iloc[2]}")
    
    # No duplicates — just find header and proceed
    if not rows_to_remove:
        print("[OK] No duplicate headers found")
        header_index = find_header_row(df_full)
        print(f"[OK] Header row index = {header_index}")

        # -------------------------------
        # FINAL EXTRA PROCESSING STEPS
        # -------------------------------
        print("\nApplying extra processing steps (remove merged + clean excel)...")

        # STEP A: Remove merged cells
        try:
            from openpyxl import load_workbook

            wb = load_workbook(vrn_file_path)
            ws = wb.active

            for merged_range in list(ws.merged_cells):
                ws.unmerge_cells(range_string=str(merged_range))

            unmerged = vrn_file_path.replace(".xlsx", "_unmerged.xlsx")
            wb.save(unmerged)
            print(f"[OK] Unmerged - {unmerged}")
        except Exception as e:
            print(f"[ERROR] Failed unmerging: {e}")
            return vrn_file_path, header_index

        # STEP B: Clean Excel (empty column removal)
        try:
            from clean_excel import remove_empty_columns
            cleaned = remove_empty_columns(unmerged, sheet_name, header_index+1)
            print(f"[OK] Cleaned - {cleaned}")
        except Exception as e:
            print(f"[ERROR] Failed cleaning: {e}")
            return vrn_file_path, header_index

        # Replace original
        os.remove(vrn_file_path)
        shutil.move(cleaned, vrn_file_path)
        os.remove(unmerged)

        print(f"[OK] Final cleaned file saved as: {vrn_file_path}")
        return vrn_file_path, header_index

    # ------------------------------
    # STEP 2: Remove duplicate headers
    # ------------------------------
    df_cleaned = df_full.drop(index=rows_to_remove).reset_index(drop=True)

    temp_file = "vrn_file_temp.xlsx"
    with pd.ExcelWriter(temp_file, engine='openpyxl') as writer:
        df_cleaned.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

    os.remove(vrn_file_path)
    os.rename(temp_file, vrn_file_path)

    print(f"[OK] Duplicate headers cleaned - Saved to {vrn_file_path}")

    # ------------------------------
    # STEP 3: Find header row
    # ------------------------------
    header_index = find_header_row(df_cleaned)
    if header_index is None:
        header_index = 11
        print("[Warning] Could not auto-detect header, using fallback index 11")
    else:
        print(f"[OK] Header index = {header_index}")

    # ------------------------------
    # FINAL EXTRA PROCESSING STEPS
    # ------------------------------
    print("\nApplying extra processing steps (remove merged + clean excel)...")

    # STEP A: Unmerge cells
    try:
        from openpyxl import load_workbook

        wb = load_workbook(vrn_file_path)
        ws = wb.active

        for merged_range in list(ws.merged_cells):
            ws.unmerge_cells(range_string=str(merged_range))

        unmerged = vrn_file_path.replace(".xlsx", "_unmerged.xlsx")
        wb.save(unmerged)
        print(f"[OK] Unmerged - {unmerged}")

    except Exception as e:
        print(f"[ERROR] Failed unmerging: {e}")
        return vrn_file_path, header_index

    # STEP B: Clean Excel (remove empty columns)
    try:
        from clean_excel import remove_empty_columns
        cleaned = remove_empty_columns(unmerged, sheet_name, header_index+1)
        print(f"[OK] Cleaned - {cleaned}")
    except Exception as e:
        print(f"[ERROR] Failed cleaning: {e}")
        return vrn_file_path, header_index

    # Replace original file with cleaned file
    os.remove(vrn_file_path)
    shutil.move(cleaned, vrn_file_path)
    os.remove(unmerged)

    print("\n" + "=" * 80)
    print("FINAL SUMMARY")
    print("=" * 80)
    print(f"Final file: {vrn_file_path}")
    print(f"Header row index: {header_index}")
    print("[OK] All processing steps completed successfully!")
    print("=" * 80)

    return vrn_file_path, header_index
