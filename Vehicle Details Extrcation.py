import time
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException, NoSuchElementException
import os

URL = "https://parivahan.gov.in/"
EXCEL_PATH = r"C:\Users\Divyesh Parmar\Downloads\Chhapar Q-2 Checkpost Run.xlsx"
WAIT_TIME = 30
ELEMENT_WAIT = 15

def get_sheet_name():
    """Dynamically determine the sheet name from the Excel file."""
    try:
        wb = load_workbook(EXCEL_PATH)
        sheet_names = wb.sheetnames
        
        preferred_names = ["Unique Veh No for weight Check", "Sheet1", "Data", "Input"]
        
        for name in preferred_names:
            if name in sheet_names:
                print(f"Using sheet: {name}")
                return name
        
        first_sheet = sheet_names[0]
        print(f"Using first sheet: {first_sheet}")
        return first_sheet
        
    except Exception as e:
        print(f"Error determining sheet name: {e}")
        return "Sheet1"

def setup_driver():
    try:
        options = webdriver.ChromeOptions()
        options.add_argument('--disable-images')
        options.add_argument('--blink-settings=imagesEnabled=false')
        options.add_argument('--disable-gpu')
        options.page_load_strategy = 'normal'
        prefs = {"profile.managed_default_content_settings.images": 2}
        options.add_experimental_option("prefs", prefs)

        driver = webdriver.Chrome(options=options)
        driver.maximize_window()
        print("Browser launched")
        return driver
    except WebDriverException as e:
        print(f"Browser setup failed: {e}")
        return None

def wait_for_page_load(driver, wait, timeout=30):
    """Wait for page to completely load."""
    try:
        print("Waiting for page to load completely...")
        wait.until(lambda driver: driver.execute_script("return document.readyState") == "complete")
        print("Page loaded completely")
        return True
    except Exception as e:
        print(f"Page load timeout: {e}")
        return False

def navigate_to_tax_page(driver, wait):
    """Navigate from parivahan.gov.in to the tax collection page."""
    try:
        print("Opening parivahan.gov.in...")
        driver.get(URL)
        
        # Wait for page to load completely
        if not wait_for_page_load(driver, wait):
            return False
        
        # Check for mobile number update popup
        popup_closed = close_mobile_popup(driver, wait)
        if popup_closed:
            print("Mobile number popup closed")
        else:
            print("No mobile number popup found")
        
        # Hover on Online Services
        print("Hovering over Online Services...")
        online_services_xpath = "//a[@id='Online' and contains(@class, 'parent-link-with-submenu')]"
        online_services = wait.until(EC.element_to_be_clickable((By.XPATH, online_services_xpath)))
        
        ActionChains(driver).move_to_element(online_services).pause(2).perform()
        time.sleep(3)
        
        # Click on Checkpost Tax
        print("Clicking Checkpost Tax...")
        checkpost_tax_xpath = "//a[@href='/en/node/579' and contains(@class, 'second-child-menu')]"
        checkpost_tax = wait.until(EC.element_to_be_clickable((By.XPATH, checkpost_tax_xpath)))
        checkpost_tax.click()
        
        # Wait for Checkpost Tax page to load
        print("Waiting for Checkpost Tax page to load...")
        checkpost_title_xpath = "//span[@class='field field--name-title field--type-string field--label-hidden' and contains(text(), 'Checkpost Tax')]"
        wait.until(EC.visibility_of_element_located((By.XPATH, checkpost_title_xpath)))
        print("Checkpost Tax page loaded successfully")
        
        # Select Tamil Nadu state
        if not select_state_tamilnadu(driver, wait):
            print("Failed to select Tamil Nadu state")
            return False
        
        # Wait for Service Name section
        print("Waiting for Service Name section...")
        service_heading_xpath = "//h3[@class='top-space' and contains(text(), 'Service Name')]"
        wait.until(EC.visibility_of_element_located((By.XPATH, service_heading_xpath)))
        print("Service Name section is visible")
        
        # Select the service
        if not select_service(driver, wait):
            print("Failed to select service")
            return False
            
        print("Successfully navigated to vehicle entry page")
        return True
        
    except Exception as e:
        print(f"Error navigating to tax page: {e}")
        return False

def select_state_tamilnadu(driver, wait):
    """Select Tamil Nadu from the state selection interface."""
    for attempt in range(3):
        try:
            print(f"Selecting Tamil Nadu state (attempt {attempt+1})")
            
            # Look for state selection element
            state_selectors = [
                "//div[contains(text(), 'Select State Name')]",
                "//span[contains(text(), 'Select State Name')]",
                "//a[contains(text(), 'Select State Name')]",
                "//button[contains(text(), 'Select State Name')]",
                "//*[contains(text(), 'Select State Name') and contains(@class, 'select')]",
                "//*[contains(text(), 'Select State Name')]"
            ]
            
            state_element = None
            for selector in state_selectors:
                try:
                    state_element = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, selector))
                    )
                    print(f"Found state selection element with: {selector}")
                    break
                except:
                    continue
            
            if not state_element:
                print("Could not find state selection element")
                return False
            
            # Click to open state selection
            driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", state_element)
            time.sleep(1)
            
            print("Clicking state selection element...")
            state_element.click()
            time.sleep(2)
            
            # Look for Tamil Nadu option
            tamilnadu_selectors = [
                "//div[contains(text(), 'TAMIL NADU')]",
                "//span[contains(text(), 'TAMIL NADU')]",
                "//a[contains(text(), 'TAMIL NADU')]",
                "//li[contains(text(), 'TAMIL NADU')]",
                "//option[contains(text(), 'TAMIL NADU')]",
                "//*[contains(text(), 'TAMIL NADU') and contains(@class, 'option')]",
                "//*[contains(text(), 'TAMIL NADU')]"
            ]
            
            tamilnadu_element = None
            for selector in tamilnadu_selectors:
                try:
                    tamilnadu_element = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, selector))
                    )
                    print(f"Found Tamil Nadu option with: {selector}")
                    break
                except:
                    continue
            
            if not tamilnadu_element:
                print("Could not find Tamil Nadu option")
                return False
            
            # Click Tamil Nadu
            driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", tamilnadu_element)
            time.sleep(1)
            
            print("Clicking Tamil Nadu...")
            tamilnadu_element.click()
            time.sleep(3)
            
            # Verify selection
            service_heading_xpath = "//h3[@class='top-space' and contains(text(), 'Service Name')]"
            try:
                WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, service_heading_xpath)))
                print("State selection successful - Service Name section found")
                return True
            except:
                print("Service Name section not found yet")
            
            try:
                updated_state_element = driver.find_element(By.XPATH, "//*[contains(text(), 'TAMIL NADU')]")
                print("State selection verified - Tamil Nadu is displayed")
                return True
            except:
                print("Tamil Nadu not displayed in state selection")
            
            print("Retrying state selection...")
            time.sleep(2)
            
        except Exception as e:
            print(f"Error selecting state on attempt {attempt+1}: {e}")
            time.sleep(2)
    
    print("Failed to select Tamil Nadu after all attempts")
    return False

def select_service(driver, wait):
    """Select service on the service selection page."""
    for attempt in range(3):
        try:
            print(f"\nSelecting service (attempt {attempt+1})")

            # Wait for service dropdown
            print("Waiting for service dropdown...")
            service_dropdown_xpath = "//span[contains(@class,'ui-selectonemenu-label') and contains(text(),'---Select Service Name---')]"
            service_dropdown = wait.until(EC.element_to_be_clickable((By.XPATH, service_dropdown_xpath)))
            
            driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", service_dropdown)
            time.sleep(1)
            
            # Click service dropdown
            print("Clicking service dropdown...")
            service_dropdown.click()
            time.sleep(2)
            
            # Select VEHICLE TAX COLLECTION (OTHER STATE)
            print("Waiting for VEHICLE TAX COLLECTION option...")
            service_option_xpath = "//li[contains(@data-label, 'VEHICLE TAX COLLECTION') and contains(@data-label, 'OTHER STATE')]"
            service_option = wait.until(EC.element_to_be_clickable((By.XPATH, service_option_xpath)))
            
            print("Selecting VEHICLE TAX COLLECTION (OTHER STATE)...")
            service_option.click()
            time.sleep(2)
            print("Selected service: VEHICLE TAX COLLECTION (OTHER STATE)")

            # Click Go button
            print("Waiting for Go button...")
            go_button_xpath = "//button[.//span[contains(text(), 'Go')]]"
            go_button = wait.until(EC.element_to_be_clickable((By.XPATH, go_button_xpath)))
            
            print("Clicking Go button...")
            go_button.click()
            
            # Wait for vehicle entry page
            print("Waiting for vehicle entry page...")
            time.sleep(5)
            
            # Verify we reached the vehicle entry page
            vehicle_input_xpath = "//input[@type='text' and @maxlength='10']"
            if element_exists(driver, By.XPATH, vehicle_input_xpath, timeout=15):
                print("Successfully loaded vehicle entry page")
                return True
            else:
                print("Vehicle entry page not loaded, retrying...")
                reload_page(driver, 3)

        except Exception as e:
            print(f"Error selecting service: {e}")
            reload_page(driver, 3)

    print("Failed to select service after multiple attempts")
    return False

def close_mobile_popup(driver, wait):
    """Close the mobile number update popup if it appears."""
    try:
        popup_text_xpath = "//span[contains(@class, 'english') and contains(text(), 'Update Your Mobile Number')]"
        popup_text = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, popup_text_xpath)))
        
        if popup_text:
            close_button_xpath = "//button[contains(@class, 'btn-close') and contains(@class, 'position-absolute')]"
            close_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, close_button_xpath)))
            close_button.click()
            print("Mobile number popup closed")
            time.sleep(2)
            return True
            
    except TimeoutException:
        return False
    except Exception as e:
        print(f"Error handling mobile popup: {e}")
        return False

def reload_page(driver, wait_seconds=3):
    """Reload page with proper waiting."""
    try:
        print("Reloading page...")
        driver.execute_cdp_cmd("Page.reload", {"ignoreCache": False})
        time.sleep(wait_seconds)
        print("Page reloaded via CDP")
    except Exception:
        driver.execute_script("location.reload()")
        time.sleep(wait_seconds)
        print("Page reloaded via JavaScript")

def element_exists(driver, by, value, timeout=10):
    """Check if element exists with proper timeout."""
    try:
        WebDriverWait(driver, timeout).until(EC.presence_of_element_located((by, value)))
        return True
    except TimeoutException:
        return False

def safe_click(driver, wait, xpath, description="element", timeout=15):
    """Safely click element with proper waits."""
    try:
        print(f"Waiting for {description} to be clickable...")
        elem = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((By.XPATH, xpath)))
        
        print(f"Scrolling to {description}...")
        driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", elem)
        time.sleep(1)
        
        print(f"Clicking {description}...")
        elem.click()
        time.sleep(1)
        print(f"Successfully clicked {description}")
        return True
        
    except TimeoutException:
        print(f"Timeout: could not find or click {description} within {timeout} seconds")
        return False
    except Exception as e:
        print(f"Error clicking {description}: {e}")
        return False

def safe_send_keys(driver, wait, xpath, keys, description="input", timeout=15):
    """Safely send keys to element with proper waits."""
    try:
        print(f"Waiting for {description} to be interactable...")
        elem = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((By.XPATH, xpath)))
        
        print(f"Clearing {description}...")
        elem.clear()
        time.sleep(0.5)
        
        print(f"Entering text in {description}...")
        elem.send_keys(keys)
        time.sleep(0.5)
        print(f"Successfully entered text in {description}")
        return True
        
    except TimeoutException:
        print(f"Timeout: could not find or interact with {description}")
        return False
    except Exception as e:
        print(f"Error sending keys to {description}: {e}")
        return False

def safe_js_click(driver, element):
    try:
        driver.execute_script("arguments[0].click();", element)
        time.sleep(1)
        return True
    except Exception as e:
        print(f"JS click failed: {e}")
        return False

def close_popup(driver, wait, timeout=5):
    """Detect and close modal popups with an OK button. Returns True if closed."""
    popup_xpath = "//div[contains(@class,'ui-dialog') and contains(@style,'display: block')]"
    try:
        print("Checking for popup...")
        popup = WebDriverWait(driver, timeout).until(EC.visibility_of_element_located((By.XPATH, popup_xpath)))
        if popup:
            ok_button_xpath = "//div[contains(@class,'ui-dialog')]//button[.//span[contains(translate(text(),'ok','OK'),'OK') or contains(translate(text(),'ok','OK'),'Ok')]]"
            try:
                ok_elem = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, ok_button_xpath)))
                try:
                    ok_elem.click()
                except Exception:
                    safe_js_click(driver, ok_elem)
                time.sleep(1)
                print("Closed popup via OK button")
                return True
            except TimeoutException:
                try:
                    any_btn = popup.find_element(By.XPATH, ".//button")
                    safe_js_click(driver, any_btn)
                    time.sleep(1)
                    print("Closed popup via fallback button click")
                    return True
                except Exception as e:
                    print(f"Could not close popup: {e}")
                    return False
    except TimeoutException:
        return False

def safe_set(df, idx, col, value):
    """Set a DataFrame cell respecting the column dtype to avoid dtype-incompatibility warnings."""
    try:
        df.at[idx, col] = "" if value is None else str(value)
    except Exception as e:
        print(f"safe_set failed for {col}: {e}")
        df.at[idx, col] = value

def get_vehicle_weight(driver):
    """Extract vehicle weight with special handling for BUS, OMNI BUS, MAXI CAB, MOTOR CAB."""
    weight = "0"
    
    try:
        # First get vehicle type to determine the logic
        vehicle_type = ""
        try:
            # Try to find vehicle type from dropdown or input
            vehicle_type_selectors = [
                "//span[contains(@class,'ui-selectonemenu-label') and contains(text(),'BUS')]",
                "//span[contains(@class,'ui-selectonemenu-label') and contains(text(),'OMNI BUS')]",
                "//span[contains(@class,'ui-selectonemenu-label') and contains(text(),'MAXI CAB')]",
                "//span[contains(@class,'ui-selectonemenu-label') and contains(text(),'MOTOR CAB')]",
                "//input[contains(@id, 'vehicle_type') and contains(@class, 'ui-state-filled')]"
            ]
            
            for selector in vehicle_type_selectors:
                try:
                    element = driver.find_element(By.XPATH, selector)
                    if element.tag_name == 'span':
                        vehicle_type = element.text.strip()
                    else:
                        vehicle_type = element.get_attribute('value') or ''
                    if vehicle_type:
                        break
                except:
                    continue
        except:
            pass
        
        print(f"Vehicle Type detected: {vehicle_type}")
        
        # Check if it's one of the special vehicle types
        special_vehicles = ["BUS", "OMNI BUS", "MAXI CAB", "MOTOR CAB"]
        is_special_vehicle = any(vehicle in vehicle_type.upper() for vehicle in special_vehicles)
        
        if is_special_vehicle:
            print(f"Special vehicle type detected: {vehicle_type}. Getting seat capacity sum.")
            # For special vehicles, get sum of seat capacity and sleeper capacity
            seat_capacity = "0"
            sleeper_capacity = "0"
            
            try:
                # Get seat capacity
                seat_cap_element = driver.find_element(By.ID, "txt_seat_cap")
                seat_capacity = seat_cap_element.get_attribute('value') or "0"
                print(f"Seat Capacity: {seat_capacity}")
            except Exception as e:
                print(f"Could not find seat capacity: {e}")
            
            try:
                # Get sleeper capacity
                sleeper_cap_element = driver.find_element(By.ID, "txt_sleeper_cap")
                sleeper_capacity = sleeper_cap_element.get_attribute('value') or "0"
                print(f"Sleeper Capacity: {sleeper_capacity}")
            except Exception as e:
                print(f"Could not find sleeper capacity: {e}")
            
            # Calculate sum
            try:
                total_capacity = int(seat_capacity) + int(sleeper_capacity)
                weight = str(total_capacity)
                print(f"Total Capacity (Weight): {weight}")
            except:
                weight = "0"
                print("Error calculating total capacity")
        
        else:
            # For regular vehicles, get the laden weight
            print("Regular vehicle type. Getting laden weight.")
            try:
                # Try to get weight from the specific element you provided
                weight_element = driver.find_element(By.ID, "txt_laden_wt")
                weight = weight_element.get_attribute('value') or "0"
                print(f"Laden Weight: {weight}")
            except Exception as e:
                print(f"Could not find laden weight by ID: {e}")
                
                # Fallback: try other weight elements
                weight_selectors = [
                    "//input[contains(@id, 'laden') and contains(@class, 'ui-state-filled')]",
                    "//input[contains(@name, 'laden') and contains(@class, 'ui-state-filled')]",
                    "//input[contains(@id, 'weight') and contains(@class, 'ui-state-filled')]",
                    "//input[contains(@name, 'weight') and contains(@class, 'ui-state-filled')]"
                ]
                
                for selector in weight_selectors:
                    try:
                        weight_element = driver.find_element(By.XPATH, selector)
                        weight = weight_element.get_attribute('value') or "0"
                        if weight != "0":
                            print(f"Found weight via fallback: {weight}")
                            break
                    except:
                        continue
        
        # Final validation
        if weight == "0" or not weight:
            print("Weight is 0 or empty, checking for any weight input...")
            # Last resort: find any input that might contain weight
            try:
                all_inputs = driver.find_elements(By.XPATH, "//input[contains(@class, 'ui-state-filled')]")
                for inp in all_inputs:
                    value = inp.get_attribute('value') or ''
                    if value and value.isdigit() and int(value) > 0:
                        # Check if this might be a weight field by looking at surrounding text
                        try:
                            parent_text = inp.find_element(By.XPATH, "./ancestor::tr[1] | ./ancestor::div[1]").text.lower()
                            if 'weight' in parent_text or 'laden' in parent_text or 'capacity' in parent_text:
                                weight = value
                                print(f"Found weight via text analysis: {weight}")
                                break
                        except:
                            continue
            except Exception as e:
                print(f"Error in final weight search: {e}")
    
    except Exception as e:
        print(f"Error in get_vehicle_weight: {e}")
    
    return weight

def extract_all_vehicle_data_fast(driver, wait):
    """OPTIMIZED: Faster data extraction using bulk element finding."""
    data = {}
    
    try:
        print("Quickly extracting vehicle data...")
        
        # OPTIMIZATION: Find all populated fields at once instead of individual searches
        all_inputs = driver.find_elements(By.XPATH, "//input[@class='ui-state-filled' or contains(@class, 'ui-inputfield')]")
        all_spans = driver.find_elements(By.XPATH, "//span[contains(@class, 'ui-selectonemenu-label') and not(contains(text(),'---Select'))]")
        
        # Process all inputs quickly
        for inp in all_inputs:
            value = inp.get_attribute('value') or ''
            if value:
                # Get surrounding context to identify field type
                try:
                    parent_text = inp.find_element(By.XPATH, "./ancestor::tr[1] | ./ancestor::div[1]").text.lower()
                    
                    if 'chassis' in parent_text:
                        data['Chassis No'] = value
                    elif 'owner' in parent_text:
                        data['Owner Name'] = value
                    elif 'mobile' in parent_text:
                        data['Mobile No'] = value
                    elif 'insurance' in parent_text:
                        data['Insurance Validity'] = value
                    elif 'fitness' in parent_text:
                        data['Fitness Validity'] = value
                    elif 'permit' in parent_text and 'authorization' in parent_text:
                        data['Permit Authorization No'] = value
                    elif 'permit' in parent_text and 'no' in parent_text:
                        data['Permit No'] = value
                    elif 'pucc' in parent_text:
                        data['PUCC Validity'] = value
                except:
                    pass
        
        # Process all spans (dropdown values) quickly
        for span in all_spans:
            text = span.text.strip()
            if text:
                try:
                    parent_text = span.find_element(By.XPATH, "./ancestor::tr[1] | ./ancestor::div[1]").text.lower()
                    
                    if 'vehicle type' in parent_text:
                        data['Vehicle Type'] = text
                    elif 'from state' in parent_text:
                        data['From State'] = text
                    elif 'permit type' in parent_text:
                        data['Permit Type'] = text
                    elif 'vehicle class' in parent_text:
                        data['vehicle class'] = text
                except:
                    pass
        
        # OPTIMIZATION: Get weight using the specialized function
        data['weight'] = get_vehicle_weight(driver)
        
        # Fill any missing fields with quick individual lookups
        field_mappings = {
            'Chassis No': ['chassis'],
            'Owner Name': ['owner'],
            'Mobile No': ['mobile'],
            'Vehicle Type': ['vehicle type'],
            'From State': ['from state'],
            'Insurance Validity': ['insurance'],
            'Fitness Validity': ['fitness'],
            'Permit Type': ['permit type'],
            'Permit No': ['permit no'],
            'Permit Authorization No': ['permit authorization'],
            'Permit Authorization Validity': ['permit authorization validity'],
            'PUCC Validity': ['pucc'],
            'vehicle class': ['vehicle class']
        }
        
        for field, keywords in field_mappings.items():
            if field not in data or not data[field]:
                for keyword in keywords:
                    try:
                        # Quick lookup by label
                        label_xpath = f"//*[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), '{keyword}')]"
                        labels = driver.find_elements(By.XPATH, label_xpath)
                        for label in labels:
                            try:
                                # Find following input or span
                                following_input = label.find_element(By.XPATH, "./following::input[1] | ./following::span[1]")
                                if following_input.tag_name == 'input':
                                    value = following_input.get_attribute('value') or ''
                                else:
                                    value = following_input.text or ''
                                
                                if value and value.strip():
                                    data[field] = value.strip()
                                    break
                            except:
                                continue
                        if field in data and data[field]:
                            break
                    except:
                        continue
        
        print("✓ Quick data extraction completed")
        for key, value in data.items():
            if value:
                print(f"  {key}: {value}")
        
    except Exception as e:
        print(f"Error during quick data extraction: {e}")
    
    return data

def save_data_to_excel(df, excel_path, sheet_name):
    """Save the entire DataFrame to Excel WITHOUT creating backup."""
    try:
        # Save directly to Excel — overwrite the file
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        print(f"Successfully saved all data to Excel: {excel_path}")
        return True

    except Exception as e:
        print(f"Error saving to Excel: {e}")
        return False

def restart_browser_and_continue(driver, current_vehicle_no, current_idx, df, SHEET_NAME):
    """Restart browser and continue from the same vehicle number."""
    print(f"\n🚨 RESTARTING BROWSER - Vehicle {current_vehicle_no} will be retried...")
    
    # Close current browser
    try:
        driver.quit()
        print("Closed current browser session")
    except:
        print("Could not properly close browser")
    
    # Wait a moment
    time.sleep(3)
    
    # Start new browser
    new_driver = setup_driver()
    if not new_driver:
        print("Could not restart browser — aborting")
        return None, False
    
    new_wait = WebDriverWait(new_driver, WAIT_TIME)
    
    # Navigate to tax page again
    print("Re-navigating to tax page...")
    if not navigate_to_tax_page(new_driver, new_wait):
        print("Could not navigate to tax page after restart — aborting")
        new_driver.quit()
        return None, False
    
    print("✓ Browser restarted successfully")
    return new_driver, new_wait

def process_single_vehicle(driver, wait, vehicle_no, idx, df, SHEET_NAME):
    """Process a single vehicle with retry logic for browser restart."""
    max_retries = 2
    retry_count = 0
    
    while retry_count <= max_retries:
        try:
            print(f"\n{'='*50}")
            print(f"Processing {idx+1}/{len(df)} — Vehicle: {vehicle_no} (Attempt {retry_count + 1})")
            print(f"{'='*50}")

            # Quick refresh
            print("Quick refreshing page...")
            driver.execute_script("location.reload()")
            time.sleep(2)
            
            # Wait for vehicle input field - THIS IS THE CRITICAL PART
            vehicle_input_xpath = "//input[@type='text' and @maxlength='10']"
            print("Waiting for Vehicle Number input to be interactable...")
            
            try:
                input_element = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, vehicle_input_xpath)))
                print("✓ Vehicle Number input found and ready")
            except TimeoutException:
                print("❌ Timeout: could not find or interact with Vehicle Number input")
                
                if retry_count < max_retries:
                    retry_count += 1
                    print(f"🔄 Attempting browser restart ({retry_count}/{max_retries})...")
                    
                    # Restart browser and continue with same vehicle
                    new_driver, new_wait = restart_browser_and_continue(driver, vehicle_no, idx, df, SHEET_NAME)
                    if new_driver:
                        driver = new_driver
                        wait = new_wait
                        continue  # Retry the same vehicle with new browser
                    else:
                        break
                else:
                    print("❌ Max retries reached for browser restart")
                    df.at[idx, "Run Status"] = "Run Complete"
                    df.at[idx, "vehicle class"] = "Error - Input Not Found"
                    safe_set(df, idx, "weight", "Error - Input Not Found")
                    return driver, wait, False
            
            # If we get here, input was found successfully
            # Quick popup check
            try:
                popup_xpath = "//div[contains(@class,'ui-dialog') and contains(@style,'display: block')]"
                popup = WebDriverWait(driver, 2).until(EC.visibility_of_element_located((By.XPATH, popup_xpath)))
                if popup:
                    ok_button_xpath = "//div[contains(@class,'ui-dialog')]//button[.//span[contains(text(),'OK') or contains(text(),'Ok')]]"
                    safe_click(driver, wait, ok_button_xpath, "OK button on popup")
            except TimeoutException:
                pass

            # Enter Vehicle Number
            try:
                input_element = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, vehicle_input_xpath)))
                driver.execute_script("arguments[0].value = arguments[1];", input_element, vehicle_no)
                print("✓ Vehicle number entered via JavaScript")
            except:
                if not safe_send_keys(driver, wait, vehicle_input_xpath, vehicle_no, "Vehicle Number input"):
                    continue

            # Click Get Details
            get_details_xpath = "//button[.//span[contains(text(), 'Get Details')]]"
            try:
                get_details_btn = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, get_details_xpath)))
                driver.execute_script("arguments[0].click();", get_details_btn)
                print("✓ Get Details clicked via JavaScript")
            except:
                if not safe_click(driver, wait, get_details_xpath, "Get Details button"):
                    continue

            print("Waiting for vehicle details...")
            time.sleep(3)

            # Check for popup or data
            popup_appeared = False
            data_appeared = False

            for attempt in range(2):
                try:
                    # Check for popup first
                    popup_xpath = "//div[contains(@class,'ui-dialog') and contains(@style,'display: block')]"
                    try:
                        popup = WebDriverWait(driver, 2).until(EC.visibility_of_element_located((By.XPATH, popup_xpath)))
                        popup_appeared = True
                        print("Popup detected - no data available")
                        break
                    except TimeoutException:
                        pass

                    # Quick data check
                    try:
                        filled_fields = driver.find_elements(By.XPATH, "//input[@class='ui-state-filled'] | //span[contains(@class,'ui-selectonemenu-label') and not(contains(text(),'---Select'))]")
                        if filled_fields:
                            data_appeared = True
                            print("✓ Vehicle details loaded")
                            break
                    except:
                        pass

                except Exception:
                    pass

                if not popup_appeared and not data_appeared:
                    time.sleep(1)

            # Handle popup case (no data)
            if popup_appeared:
                ok_button_xpath = "//div[contains(@class,'ui-dialog')]//button[.//span[contains(text(),'OK') or contains(text(),'Ok')]]"
                safe_click(driver, wait, ok_button_xpath, "OK button on popup")
                df.at[idx, "Run Status"] = "Run Complete"
                df.at[idx, "vehicle class"] = "N/A"
                safe_set(df, idx, "weight", "N/A")
                print("✓ Marked as completed (no data available)")

            # Handle data case
            elif data_appeared:
                vehicle_data = extract_all_vehicle_data_fast(driver, wait)
                
                for field, value in vehicle_data.items():
                    if field in df.columns:
                        df.at[idx, field] = value
                
                df.at[idx, "Run Status"] = "Run Complete"
                print("✓ Quick data extraction completed")

            else:
                print("No data loaded after quick attempts")
                df.at[idx, "Run Status"] = "Run Complete"
                df.at[idx, "vehicle class"] = "No Data"
                safe_set(df, idx, "weight", "No Data")

            # If we get here, processing was successful
            return driver, wait, True

        except Exception as e:
            print(f"✗ Error processing {vehicle_no}: {e}")
            
            if retry_count < max_retries:
                retry_count += 1
                print(f"🔄 Attempting browser restart due to error ({retry_count}/{max_retries})...")
                
                # Restart browser and continue with same vehicle
                new_driver, new_wait = restart_browser_and_continue(driver, vehicle_no, idx, df, SHEET_NAME)
                if new_driver:
                    driver = new_driver
                    wait = new_wait
                    continue  # Retry the same vehicle with new browser
                else:
                    break
            else:
                print("❌ Max retries reached after error")
                df.at[idx, "Run Status"] = "Run Complete"
                df.at[idx, "vehicle class"] = "Error"
                safe_set(df, idx, "weight", "Error")
                return driver, wait, False
    
    # If we exhaust all retries
    df.at[idx, "Run Status"] = "Run Complete"
    df.at[idx, "vehicle class"] = "Error - Max Retries"
    safe_set(df, idx, "weight", "Error - Max Retries")
    return driver, wait, False

def main():
    SHEET_NAME = get_sheet_name()
    print(f"Processing sheet: {SHEET_NAME}")
    
    # Load Excel into pandas
    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME, dtype=str)
        print(f"Loaded {len(df)} vehicles from Excel")
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    # Define required columns
    required_columns = [
        "Unique Vehicle Number", "Chassis No", "Owner Name", "Mobile No", 
        "Vehicle Type", "From State", "Insurance Validity", "Fitness Validity",
        "Permit Type", "Permit No", "Permit Authorization No", 
        "Permit Authorization Validity", "PUCC Validity", "vehicle class", 
        "weight", "Run Status"
    ]
    
    # Add missing columns
    for col in required_columns:
        if col not in df.columns:
            df[col] = ""
            print(f"Added missing column: {col}")
    
    # Start browser
    driver = setup_driver()
    if not driver:
        print("Could not start browser — aborting")
        return

    wait = WebDriverWait(driver, WAIT_TIME)
    
    # Navigate to tax page
    print("Starting navigation to tax page...")
    if not navigate_to_tax_page(driver, wait):
        print("Could not navigate to tax page — aborting")
        driver.quit()
        return

    start_time = time.perf_counter()
    total = len(df)
    processed_count = 0

    for idx, row in df.iterrows():
        vehicle_no = str(row["Vehicle Number"]).strip()
        
        # Skip if already completed
        current_status = str(row.get("Run Status", "")).strip()
        if current_status.lower() == "run complete":
            print(f"✓ Already completed - skipping vehicle {idx+1}")
            processed_count += 1
            continue

        # Process single vehicle with retry logic
        driver, wait, success = process_single_vehicle(driver, wait, vehicle_no, idx, df, SHEET_NAME)
        
        if success:
            processed_count += 1
        
        # SAVE AFTER EACH VEHICLE
        print(f"💾 Saving data to Excel... (Progress: {processed_count}/{total})")
        if save_data_to_excel(df, EXCEL_PATH, SHEET_NAME):
            print("✓ Successfully saved to Excel")
        else:
            print("✗ Failed to save to Excel")

        # Short wait before next vehicle
        time.sleep(1)

    # Final summary
    print(f"\n{'='*50}")
    print(f"PROCESSING COMPLETE")
    print(f"Total vehicles: {total}")
    print(f"Processed: {processed_count}")
    print(f"Skipped (already completed): {total - processed_count}")
    print(f"{'='*50}")

    # Final save
    try:
        save_data_to_excel(df, EXCEL_PATH, SHEET_NAME)
        print("✓ Final save completed successfully")
    except Exception as e:
        print(f"✗ Final save failed: {e}")

    # Calculate and display execution time
    elapsed = time.perf_counter() - start_time
    hrs = int(elapsed // 3600)
    mins = int((elapsed % 3600) // 60)
    secs = int(elapsed % 60)
    print(f"Total execution time: {hrs:02d}:{mins:02d}:{secs:02d} (HH:MM:SS)")

    driver.quit()

if __name__ == "__main__":
    main()